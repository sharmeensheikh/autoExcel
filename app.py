import openpyxl as xl

wb = xl.load_workbook('prices.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
header = sheet['d1']
header.value = 'Discounted Price'

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    changed_val = cell.value * 0.5
    changed_val_cell = sheet.cell(row, 4)
    changed_val_cell.value = changed_val


wb.save('prices_changed.xlsx')
print("Success!")